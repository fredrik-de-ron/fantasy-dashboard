import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

with open("spelare_komplett.json", encoding="utf-8") as f:
    spelare = json.load(f)

with open("kombinerad.json", encoding="utf-8") as f:
    kombinerad = json.load(f)

wb = openpyxl.Workbook()

BLA_MRK = "1F4E79"
GRN_MRK = "375623"
VIT     = "FFFFFF"
GRA_LJU = "F2F2F2"

POS_FARG = {
    "Målvakt":    "DDEBF7",
    "Försvarare": "E2EFDA",
    "Mittfältare":"FFF2CC",
    "Anfallare":  "FCE4D6",
}

def rubrik_rad(ws, rubriker, farg=BLA_MRK):
    for kol, text in enumerate(rubriker, 1):
        c = ws.cell(row=1, column=kol, value=text)
        c.font = Font(bold=True, color=VIT, size=10)
        c.fill = PatternFill("solid", fgColor=farg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

def farg_rad(ws, rad, pos, n_kol):
    farg = POS_FARG.get(pos)
    if farg:
        for kol in range(1, n_kol + 1):
            ws.cell(row=rad, column=kol).fill = PatternFill("solid", fgColor=farg)
    elif rad % 2 == 0:
        for kol in range(1, n_kol + 1):
            ws.cell(row=rad, column=kol).fill = PatternFill("solid", fgColor=GRA_LJU)

def set_bredder(ws, bredder):
    for i, b in enumerate(bredder, 1):
        ws.column_dimensions[get_column_letter(i)].width = b

def safe_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return v

# =====================
# FLIK 1: Spelaröversikt
# =====================
ws1 = wb.active
ws1.title = "Spelare"

rub1 = [
    "ID", "Namn", "Lag", "Position", "Status", "Chans spela %",
    "Pris", "Prisänd omgång", "Prisänd säsong",
    "Ägarskap %", "Transfers in", "Transfers ut",
    "Form", "Värde form", "Värde säsong", "EP nästa", "EP denna",
    "Poäng total", "Poäng/match", "Poäng omgång", "Minuter",
    "Mål", "Assist", "Gula", "Röda", "Nollor",
    "Insläppta", "Räddningar", "Straffräddningar",
    "Missade straff", "Självmål", "Avgörande mål",
    "Nyckelpass", "Def. aktioner", "Off. bonus", "Def. bonus"
]

rubrik_rad(ws1, rub1)

for rad, s in enumerate(spelare, 2):
    varden = [
        s["id"], s["namn"], s["lag"], s["position"],
        s["status"], s["chans_spela_nästa"],
        s["pris"], s["prisändring_omgång"], s["prisändring_säsong"],
        s["agarskap"], s["transfers_in_totalt"], s["transfers_out_totalt"],
        s["form"], s["varde_form"], s["varde_säsong"],
        s["ep_nästa"], s["ep_denna"],
        s["poang_total"], s["poang_per_match"], s["poang_omgång"],
        s["minuter"], s["mal"], s["assist"],
        s["gula_kort"], s["roda_kort"], s["nollor"],
        s["inslappta_mal"], s["radningar"], s["straffradningar"],
        s["missade_straff"], s["sjalvmal"], s["avgörande_mal"],
        s["nyckelpassningar"], s["defensiva_aktioner"],
        s["offensiv_bonus"], s["defensiv_bonus"]
    ]
    for kol, v in enumerate(varden, 1):
        c = ws1.cell(row=rad, column=kol, value=v)
        c.alignment = Alignment(horizontal="center")
    farg_rad(ws1, rad, s["position"], len(rub1))

ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:{get_column_letter(len(rub1))}1"
set_bredder(ws1, [5,16,16,12,7,12,6,12,12,10,11,11,7,10,11,9,9,11,11,11,8,6,7,6,6,7,9,10,14,12,8,13,10,13,10,10])

print("Flik 1 klar")

# =====================
# FLIK 2: Per omgång
# =====================
ws2 = wb.create_sheet("Per omgång")

rub2 = [
    "Namn", "Lag", "Position", "Omgång",
    "Poäng", "Minuter", "Mål", "Assist",
    "Gula", "Röda", "Nollor", "Insläppta",
    "Räddningar", "Straffräddningar", "Missade straff",
    "Självmål", "Avgörande mål", "Nyckelpass",
    "Def. aktioner", "Off. bonus", "Def. bonus"
]

rubrik_rad(ws2, rub2)

rad2 = 2
for s in spelare:
    for gw in s.get("omgangar", []):
        varden2 = [
            s["namn"], s["lag"], s["position"], gw["omgang"],
            gw["poang"], gw["minuter"], gw["mal"], gw["assist"],
            gw["gula_kort"], gw["roda_kort"], gw["nollor"],
            gw["inslappta_mal"], gw["radningar"], gw["straffradningar"],
            gw["missade_straff"], gw["sjalvmal"], gw["avgörande_mal"],
            gw["nyckelpassningar"], gw["defensiva_aktioner"],
            gw["offensiv_bonus"], gw["defensiv_bonus"]
        ]
        for kol, v in enumerate(varden2, 1):
            c = ws2.cell(row=rad2, column=kol, value=v)
            c.alignment = Alignment(horizontal="center")
        farg_rad(ws2, rad2, s["position"], len(rub2))
        rad2 += 1

ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:{get_column_letter(len(rub2))}1"
set_bredder(ws2, [16,16,12,8,7,8,6,7,6,6,7,9,10,14,12,8,13,10,13,10,10])

print("Flik 2 klar")

# =====================
# FLIK 3: Prisrörelser
# =====================
ws3 = wb.create_sheet("Prisrörelser")

rub3 = [
    "Namn", "Lag", "Position", "Pris",
    "Ändring omgång", "Ändring säsong",
    "Transfers in omgång", "Transfers ut omgång",
    "Ägarskap %", "EP nästa", "Form"
]

rubrik_rad(ws3, rub3, farg=GRN_MRK)

med_andring = [s for s in spelare if s["prisändring_omgång"] != 0]
sorterad3 = sorted(med_andring, key=lambda x: x["prisändring_omgång"], reverse=True)

rad3 = 2
for s in sorterad3:
    varden3 = [
        s["namn"], s["lag"], s["position"], s["pris"],
        s["prisändring_omgång"], s["prisändring_säsong"],
        s["transfers_in_omgång"], s["transfers_out_omgång"],
        s["agarskap"], s["ep_nästa"], s["form"]
    ]
    for kol, v in enumerate(varden3, 1):
        c = ws3.cell(row=rad3, column=kol, value=v)
        c.alignment = Alignment(horizontal="center")
        if s["prisändring_omgång"] > 0:
            c.fill = PatternFill("solid", fgColor="E2EFDA")
        else:
            c.fill = PatternFill("solid", fgColor="FCE4D6")
    rad3 += 1

ws3.freeze_panes = "A2"
ws3.auto_filter.ref = f"A1:{get_column_letter(len(rub3))}1"
set_bredder(ws3, [16,16,12,7,13,13,18,18,10,9,7])

print("Flik 3 klar")

# =====================
# FLIK 4: Kaptenstips
# =====================
ws4 = wb.create_sheet("Kaptenstips")

rub4 = [
    "Rank", "Namn", "Lag", "Position", "Pris",
    "EP nästa", "Form", "Poäng/match", "Ägarskap %",
    "Minuter", "Mål", "Assist", "Chans spela %"
]

rubrik_rad(ws4, rub4, farg="7B2D8B")

kandidater = [s for s in spelare if (s["chans_spela_nästa"] or 0) >= 75]
sorterad4 = sorted(kandidater, key=lambda x: x["ep_nästa"], reverse=True)

for rad4, s in enumerate(sorterad4[:20], 2):
    varden4 = [
        rad4 - 1, s["namn"], s["lag"], s["position"], s["pris"],
        s["ep_nästa"], s["form"], s["poang_per_match"], s["agarskap"],
        s["minuter"], s["mal"], s["assist"], s["chans_spela_nästa"]
    ]
    for kol, v in enumerate(varden4, 1):
        c = ws4.cell(row=rad4, column=kol, value=v)
        c.alignment = Alignment(horizontal="center")
        if rad4 == 2:
            c.fill = PatternFill("solid", fgColor="F4E6F7")

ws4.freeze_panes = "A2"
set_bredder(ws4, [6,16,16,12,7,9,7,12,10,8,6,7,12])

print("Flik 4 klar")

# =====================
# FLIK 5: Avancerad statistik (FotMob)
# =====================
ws5 = wb.create_sheet("Avancerad statistik")

rub5 = [
    "Namn", "Lag", "Position",
    "xG", "xG p90", "xGOT", "xGOT p90",
    "xA", "xA p90",
    "Skott", "Skott p90", "Skott på mål", "Skott på mål p90",
    "Chanser skapade", "Chanser p90",
    "Dribblingar", "Dribb %",
    "Dueller vunna", "Dueller %",
    "Luftdueller", "Luftd %",
    "Bollkontakter", "Ber. straffomr p90",
    "Ingripanden", "Tacklingar",
    "Återvinningar", "Rensningar",
    "xG emot p90", "FotMob-betyg"
]

rubrik_rad(ws5, rub5, farg="7B2D8B")

rad5 = 2
for s in kombinerad:
    fm = s.get("fotmob") or {}
    varden5 = [
        s["namn"], s["lag"], s["position"],
        safe_float(fm.get("expected_goals")),
        safe_float(fm.get("expected_goals_p90")),
        safe_float(fm.get("expected_goals_on_target")),
        safe_float(fm.get("expected_goals_on_target_p90")),
        safe_float(fm.get("expected_assists")),
        safe_float(fm.get("expected_assists_p90")),
        safe_float(fm.get("shots")),
        safe_float(fm.get("shots_p90")),
        safe_float(fm.get("ShotsOnTarget")),
        safe_float(fm.get("ShotsOnTarget_p90")),
        safe_float(fm.get("chances_created")),
        safe_float(fm.get("chances_created_p90")),
        safe_float(fm.get("dribbles_succeeded")),
        safe_float(fm.get("won_contest_subtitle")),
        safe_float(fm.get("duel_won")),
        safe_float(fm.get("duel_won_percent")),
        safe_float(fm.get("aerials_won")),
        safe_float(fm.get("aerials_won_percent")),
        safe_float(fm.get("touches")),
        safe_float(fm.get("touches_opp_box_p90")),
        safe_float(fm.get("interceptions")),
        safe_float(fm.get("matchstats.headers.tackles")),
        safe_float(fm.get("recoveries")),
        safe_float(fm.get("clearances")),
        safe_float(fm.get("expected_goals_against_while_on_pitch_p90")),
        safe_float(fm.get("rating")),
    ]
    for kol, v in enumerate(varden5, 1):
        c = ws5.cell(row=rad5, column=kol, value=v)
        c.alignment = Alignment(horizontal="center")
        if rad5 % 2 == 0:
            c.fill = PatternFill("solid", fgColor=GRA_LJU)
    rad5 += 1

ws5.freeze_panes = "A2"
ws5.auto_filter.ref = f"A1:{get_column_letter(len(rub5))}1"
set_bredder(ws5, [18,16,12,7,8,7,8,7,8,7,9,11,12,14,12,11,9,14,11,11,9,13,15,11,11,12,11,13,12])

print("Flik 5 klar")

# =====================
# Spara
# =====================
filnamn = "fantasy_allsvenskan.xlsx"
wb.save(filnamn)

print(f"\nKlar! Sparade {filnamn}")
print(f"  Flik 1 — Spelare:             {len(spelare)} rader")
print(f"  Flik 2 — Per omgång:          {rad2 - 2} rader")
print(f"  Flik 3 — Prisrörelser:        {rad3 - 2} rader")
print(f"  Flik 4 — Kaptenstips:         topp 20 för nästa omgång")
print(f"  Flik 5 — Avancerad statistik: {rad5 - 2} rader med FotMob-data")